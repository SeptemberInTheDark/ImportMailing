from __future__ import annotations

import logging
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from mailings.models import MailingMessage

logger = logging.getLogger(__name__)


REQUIRED_COLUMNS = ["external_id", "user_id", "email", "subject", "message"]


@dataclass(frozen=True)
class ImportStats:
    processed_rows: int = 0
    created: int = 0
    skipped: int = 0
    invalid_rows: int = 0


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


class Command(BaseCommand):
    help = "Импортирует рассылки из XLSX и имитирует отправку писем (лог + задержка)."

    def add_arguments(self, parser) -> None:
        parser.add_argument("xlsx_path", type=str, help="Путь к файлу XLSX")
        parser.add_argument(
            "--chunk-size",
            type=int,
            default=500,
            help="Размер пачки для работы с БД (по умолчанию: 500)",
        )
        parser.add_argument(
            "--send-delay-ms",
            type=int,
            default=50,
            help="Задержка имитации отправки одного письма в мс (по умолчанию: 50)",
        )

    def handle(self, *args, **options) -> None:
        xlsx_path = Path(options["xlsx_path"]).expanduser()
        chunk_size = int(options["chunk_size"])
        send_delay_ms = int(options["send_delay_ms"])

        if chunk_size <= 0:
            raise CommandError("Параметр --chunk-size должен быть больше нуля.")
        if send_delay_ms < 0:
            raise CommandError("Параметр --send-delay-ms не может быть отрицательным.")

        if not xlsx_path.exists() or not xlsx_path.is_file():
            raise CommandError(f"Файл не найден: {xlsx_path}")

        rows = self._read_xlsx_rows(xlsx_path)
        stats = ImportStats()

        buffer: list[dict[str, str]] = []
        for row in rows:
            stats = ImportStats(
                processed_rows=stats.processed_rows + 1,
                created=stats.created,
                skipped=stats.skipped,
                invalid_rows=stats.invalid_rows,
            )
            buffer.append(row)
            if len(buffer) >= chunk_size:
                stats = self._process_chunk(buffer, stats, send_delay_ms)
                buffer = []

        if buffer:
            stats = self._process_chunk(buffer, stats, send_delay_ms)

        self.stdout.write(
            self.style.SUCCESS(
                "\n".join(
                    [
                        "Результат импорта:",
                        f"- обработано строк: {stats.processed_rows}",
                        f"- создано записей: {stats.created}",
                        f"- пропущено записей: {stats.skipped}",
                        f"- ошибочных строк: {stats.invalid_rows}",
                    ]
                )
            )
        )

    def _read_xlsx_rows(self, path: Path) -> Iterable[dict[str, str]]:
        try:
            from openpyxl import load_workbook
        except Exception as exc:  # pragma: no cover
            raise CommandError(
                "Не удалось импортировать openpyxl. Установите зависимости из requirements.txt."
            ) from exc

        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        ws = wb.active

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise CommandError("Файл пустой: отсутствует строка заголовков.")

        headers = [_normalize_header(v) for v in header_row]
        missing = [c for c in REQUIRED_COLUMNS if c not in headers]
        if missing:
            raise CommandError(f"В файле отсутствуют обязательные колонки: {', '.join(missing)}")

        idx = {name: headers.index(name) for name in REQUIRED_COLUMNS}

        for row in ws.iter_rows(min_row=2, values_only=True):
            yield {
                "external_id": _cell_to_str(row[idx["external_id"]] if idx["external_id"] < len(row) else ""),
                "user_id": _cell_to_str(row[idx["user_id"]] if idx["user_id"] < len(row) else ""),
                "email": _cell_to_str(row[idx["email"]] if idx["email"] < len(row) else ""),
                "subject": _cell_to_str(row[idx["subject"]] if idx["subject"] < len(row) else ""),
                "message": _cell_to_str(row[idx["message"]] if idx["message"] < len(row) else ""),
            }

    def _process_chunk(self, chunk: list[dict[str, str]], stats: ImportStats, send_delay_ms: int) -> ImportStats:
        external_ids = [r["external_id"] for r in chunk if r.get("external_id")]

        existing = set(
            MailingMessage.objects.filter(external_id__in=external_ids).values_list("external_id", flat=True)
        )

        to_create: list[MailingMessage] = []
        invalid = 0
        skipped = 0
        seen_in_file_chunk: set[str] = set()

        for r in chunk:
            if not all(r.get(k) for k in REQUIRED_COLUMNS):
                invalid += 1
                logger.warning("Строка пропущена: не заполнены обязательные поля. external_id=%r", r.get("external_id"))
                continue

            if r["external_id"] in existing:
                skipped += 1
                logger.info("Запись пропущена: уже импортирована. external_id=%s", r["external_id"])
                continue

            if r["external_id"] in seen_in_file_chunk:
                skipped += 1
                logger.info("Запись пропущена: дубликат в файле. external_id=%s", r["external_id"])
                continue
            seen_in_file_chunk.add(r["external_id"])

            to_create.append(
                MailingMessage(
                    external_id=r["external_id"],
                    user_id=r["user_id"],
                    email=r["email"],
                    subject=r["subject"],
                    message=r["message"],
                )
            )

        created = 0
        if to_create:
            created_external_ids = [m.external_id for m in to_create]
            with transaction.atomic():
                MailingMessage.objects.bulk_create(to_create, batch_size=500)
            created = len(to_create)

            now = timezone.now()
            for msg in MailingMessage.objects.filter(external_id__in=created_external_ids).only(
                "external_id", "user_id", "email", "subject"
            ):
                logger.info(
                    "Имитация отправки письма: external_id=%s user_id=%s email=%s тема=%r",
                    msg.external_id,
                    msg.user_id,
                    msg.email,
                    msg.subject,
                )
                if send_delay_ms:
                    time.sleep(send_delay_ms / 1000.0)
            MailingMessage.objects.filter(external_id__in=created_external_ids).update(sent_at=now)

        return ImportStats(
            processed_rows=stats.processed_rows,
            created=stats.created + created,
            skipped=stats.skipped + skipped,
            invalid_rows=stats.invalid_rows + invalid,
        )

